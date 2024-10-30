import pandas as pd
from datetime import datetime
from openpyxl import Workbook

def calculate_age(birthdate: str) -> int:
    """Розрахунок повного віку на основі дати народження."""
    birth_date = datetime.strptime(birthdate, "%d-%m-%Y")
    today = datetime.today()
    return today.year - birth_date.year - ((today.month, today.day) < (birth_date.month, birth_date.day))

def categorize_age(age: int) -> str:
    """Категоризація віку для різних аркушів."""
    if age < 18:
        return "younger_18"
    elif 18 <= age <= 45:
        return "18-45"
    elif 45 < age <= 70:
        return "45-70"
    else:
        return "older_70"

try:
    # Читаємо дані з CSV файлу
    try:
        data = pd.read_csv("people_data.csv")
    except FileNotFoundError:
        print("Помилка: Файл people_data.csv не знайдено.")
        exit()
    except Exception as e:
        print(f"Помилка відкриття файлу CSV: {e}")
        exit()

    # Додаємо стовпець з віком
    data['Вік'] = data['Дата народження'].apply(calculate_age)

    # Ініціалізуємо новий XLSX файл з аркушами
    workbook = Workbook()
    sheets = {
        "all": workbook.active,
        "younger_18": workbook.create_sheet("younger_18"),
        "18-45": workbook.create_sheet("18-45"),
        "45-70": workbook.create_sheet("45-70"),
        "older_70": workbook.create_sheet("older_70")
    }
    sheets["all"].title = "all"

    # Записуємо всі дані в аркуш "all"
    all_columns = ["№"] + list(data.columns)
    sheets["all"].append(all_columns)
    for index, row in data.iterrows():
        sheets["all"].append([index + 1] + list(row))

    # Записуємо дані в інші аркуші відповідно до вікових категорій
    category_columns = ["№", "Прізвище", "Ім’я", "По батькові", "Дата народження", "Вік"]
    for category, sheet in sheets.items():
        if category != "all":
            sheet.append(category_columns)
            category_data = data[data['Вік'].apply(categorize_age) == category]
            for index, row in category_data.iterrows():
                sheet.append([index + 1, row["Прізвище"], row["Ім’я"], row["По батькові"], row["Дата народження"], row["Вік"]])

    # Зберігаємо в XLSX файл
    try:
        workbook.save("categorized_people_data.xlsx")
        print("Ok")
    except Exception as e:
        print(f"Помилка при створенні XLSX файлу: {e}")

except Exception as e:
    print(f"Виникла помилка: {e}")